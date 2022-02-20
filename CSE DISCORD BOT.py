import discord
import random
import time
from user import *
from attendance import *
from discord.ext import commands, tasks
from openpyxl import load_workbook, Workbook

token = open("token", "r").readline() 
intents = discord.Intents.default()
intents.members = True

global client
global _stat
_stat=0
box=[]
prefix = "-"
client = commands.Bot(command_prefix=prefix,intents=intents, case_insensitive=True)

@client.event
async def on_ready():
    print("Ready")
    await client.change_presence(status=discord.Status.online, activity=discord.Activity(type=discord.ActivityType.listening, name="여러분의 의견을 "))
    for guild in client.guilds:
        for channel in guild.voice_channels:
            for member in channel.members:
                print(member.id)
                result = await client.fetch_user(member.id)
                print(result)

#출석기능, 구현중#
@commands.has_role("학생회")
@client.command(name='출석')
async def attendance(ctx):
    await ctx.send(f'출석 멤버')
    for guild in client.guilds:
        for channel in guild.voice_channels:
            for member in channel.members:
                result = await client.fetch_user(member.id)
                await ctx.send(f'{result}')

#TESTING ONLY#
@commands.has_role("학생회")
@client.command(name='주사위')
async def roll(ctx, number: int):
    await ctx.send(f'주사위를 굴려 {random.randint(1,number)}이(가) 나왔습니다. (1-{number})')
@roll.error
async def roll_error(ctx, error):
    await ctx.send("명령어 오류")

# 추첨권 조회 (DM으로 개개인 갯수 보내줌)
@client.command(name='check')
async def botdm(ctx):
    #await ctx.message.author.send('hi my name is *bot name here* and i am a bot!')
    userExistance, userRow = checkUser(ctx.author.name, ctx.author.id)
    if userExistance==True:
        print("DB에서 ", ctx.author.name, "을 찾았습니다.")
        result = getMoney(ctx.author.name,userRow)
        await ctx.message.author.send(f'{ctx.author.name}님에게는 추첨권이 {result}장 있습니다.')
    else:
        print("DB에서 ", ctx.author.name, "을 찾을 수 없습니다")
        print("")
        await ctx.message.author.send("오류가 발생했습니다. 학생회에게 문의해 주세요!")

@commands.has_role("학생회")
@client.command(name='회원가입')
async def 회원가입(ctx):
    for guild in client.guilds:
        for member in guild.members:
            print(member.name,member.id)
            print("회원가입이 가능한지 확인합니다.")
            userExistance,temp = checkUser(member.name, member.id)
            if userExistance==True:
                print("DB에서 ", member.name, "을 찾았습니다.")
                print("------------------------------\n")
                await ctx.send(f'{member.name}님은 이미 가입되었습니다.')
            else:
                print("DB에서 ", member.name, "을 찾을 수 없습니다")
                print("")
                Signup(member.name, member.id)
                print("회원가입이 완료되었습니다.")
                print("------------------------------\n")
                await ctx.send(f'{member.name}님의 가입이 완료되었습니다.')

@commands.has_role("학생회")
@client.command(name='DB')
async def getDB(ctx):
    with open("userDB.xlsx", "rb") as file:
        await ctx.send("전체 DB : ", file=discord.File(file, "userDB.xlsx"))

@commands.has_role("학생회")
@client.command(name='추가')
async def add(ctx, user: discord.User, money:int):
    userExistance, userRow = checkUser(user.name, user.id)
    if userExistance==True:
        print("DB에서 ", user.name, "을 찾았습니다.")
        result= addMoney(user.name,userRow,money)
        await ctx.send(f'{user.name}에게 추첨권을 {money}장 지급하여 {result}장이 되었습니다.')
        if result > 14:
            await ctx.send(f'경고: {user.name}의 추첨권이 14장을 초과합니다.')
    else:
        print("DB에서 ", ctx.author.name, "을 찾을 수 없습니다")
        print("")
        await ctx.send("해당 회원은 없습니다.")

@commands.has_role("학생회")
@client.command(name='제거')
async def _sub(ctx, user: discord.User, money:int):
    userExistance, userRow = checkUser(user.name, user.id)
    if userExistance==True:
        print("DB에서 ", user.name, "을 찾았습니다.")
        result= subMoney(user.name,userRow,money)
        await ctx.send(f'{user.name}에게 추첨권을 {money}장 제거하여 {result}장이 되었습니다.')
        if result < 0:
            await ctx.send(f'경고: {user.name}의 추첨권이 0장 미만입니다.')
    else:
        print("DB에서 ", ctx.author.name, "을 찾을 수 없습니다")
        print("")
        await ctx.send("해당 회원은 없습니다.")

@client.command(name='조회')
async def infoMoney(ctx, user: discord.User):
    userExistance, userRow = checkUser(user.name, user.id)
    if userExistance==True:
        print("DB에서 ", user.name, "을 찾았습니다.")
        result = getMoney(user.name,userRow)
        await ctx.send(f'{user.name}에게는 추첨권이 {result}장 있습니다.')
    else:
        print("DB에서 ", ctx.author.name, "을 찾을 수 없습니다")
        print("")
        await ctx.send("해당 회원은 등록되지 않았습니다.")

#====================================================================================================#
# 추첨 기능 
@commands.has_role("추첨자")
@client.command(name='추첨준비')
async def shuffle(ctx):   #box = 추첨을 위해 미리 만들어둔 리스트 
    now=time.localtime()  
    with open("userDB.xlsx", "rb") as file:
        await ctx.message.author.send(f'{now.tm_hour}:{now.tm_min}:{now.tm_sec} DB백업:', file=discord.File(file, "userDB.xlsx")) #DB백업   
    loadFile()
    for row in range(2,ws.max_row+1):
        for _ in range(int(ws.cell(row,3).value)):
            box.append(ws.cell(row,1).value)  # DB에서 추첨권 개수만큼 사용자의 이름을 List에 추가 
    for _ in range(50):
        random.shuffle(box)  # random 함수 이용해서 List 섞음
    await ctx.send("추첨이 준비되었습니다.")

@commands.has_role("추첨자")
@client.command(name='추첨')
async def raffle(ctx):
    if box: 
        await ctx.send(f'당첨자 : {box[0]}') #0번째 인덱스를 당첨자로 선택
        temp=box[0]
        #print(temp)
        while temp in box:  
            box.remove(temp)  #뽑힌 사람의 이름을 리스트에서 제거
    else:
        await ctx.send(f'더 이상 추첨권을 가진 사람이 없습니다.')
    
@commands.has_role("추첨자")
@client.command(name='추첨초기화')
async def shuffle(ctx):
    box.clear() #리스트 초기화
    await ctx.send("추첨 초기화가 완료되었습니다.")
#=====================================================================================================#
isStart=False
duplicateList=[]
#wxb = load_workbook("attendance.xlsx")
#wxs = wb.active

@commands.has_role("학생회")
@client.command(name='출석시작')
async def getMember(ctx):
    for guild in client.guilds:
        for member in guild.members:
            print(member.name,member.id)

#DEBUG ONLY# 
@commands.has_role("학생회")
@client.command(name='member')
async def getMember(ctx):
    for guild in client.guilds:
        for member in guild.members:
            print(member.name,member.id)

@client.event
async def on_voice_state_update(member, before, after):
  if not before.channel and after.channel and member.id == 295826228114489344:
    channel = client.get_channel(938628071794307143)
    await channel.send('BOB IS HERE')

@client.event
async def on_command_error(ctx, error):
    if isinstance(error, commands.CommandNotFound):
    	await ctx.send("명령어를 찾지 못했습니다")
   
client.run(token) #토큰